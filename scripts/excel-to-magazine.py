#!/usr/bin/env python3
"""
excel-to-magazine.py  —  LOCAL-ONLY import (never runs in CI).

Reads ligne_rouge_tours_dark_tourism_seo_machine_v4.xlsx and produces the two
files the site renders from:

  src/data/magazine-pages.json   cleaned per-page briefs + assigned publishAt dates
  src/data/magazine-links.json   internal-link graph (remapped + deduped), keyed by source URL

It is the ONLY place that touches the Excel. Re-run with `npm run mag:import`
ONLY when the workbook changes. It NEVER reads or writes content/magazine/*.md.

What it fixes vs. the raw plan (see CLAUDE.md "Excel critique"):
  - canonicalises grammatically-broken / keyword-stuffed leaf slugs, type-aware,
    applied consistently to every URL reference (pages, parents, link graph);
  - remaps the dangling hubs  /magazine/ -> /dark-tourism/  and
    /safety-ethics/ -> /dark-tourism/safety-ethics/ , and synthesises the
    safety-ethics hub page (642 links point at it but it exists nowhere);
  - keeps the plan's Meta Title/Description only as informational fallback
    (522 plan titles are truncated with "…"); the live metas are hand-written
    in each .md frontmatter and uniqueness-enforced by the validator;
  - assigns a publishAt ramp date to every page (structural hubs first).
"""

import json
import re
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "ligne_rouge_tours_dark_tourism_seo_machine_v4.xlsx"
OUT_PAGES = ROOT / "src" / "data" / "magazine-pages.json"
OUT_LINKS = ROOT / "src" / "data" / "magazine-links.json"

MAGAZINE_ROOT = "/dark-tourism/"
SAFETY_HUB = "/dark-tourism/safety-ethics/"

# ---- publish ramp -----------------------------------------------------------
# A live page only appears once its .md exists AND publishAt <= build date.
# Slow ramp for a low-authority domain: 1/day for the first two weeks, then 3/day.
START_DATE = date(2026, 6, 18)   # first publish slot
RAMP_DAYS = 14
RAMP_PER_DAY = 1
CRUISE_PER_DAY = 3

# Structural pages publish before their spokes so the silo/breadcrumbs light up first.
TYPE_RANK = {
    "Main Hub": 0,
    "Cluster Hub": 1,
    "Safety Hub": 1,
    "Site Hub": 2,
}
PRIORITY_RANK = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}

HUB_TYPES = {"Main Hub", "Cluster Hub", "Site Hub", "Safety Hub"}

# Type-aware canonical leaf slug. None = keep the entity/cluster slug as-is.
# (The full path stays unique because each site hub has at most one article per type;
#  itineraries keep their day-count, cross-cluster features keep their own leaf.)
LEAF_CANON = {
    "Top List": "things-to-know-before-visiting",
    "History Explainer": "what-happened-here",
    "Mistakes": "mistakes-to-avoid",
    "Tickets / Access": "tickets-and-access",
    "Safety / Access": "is-it-safe-to-visit",
    "Ethics": "how-to-visit-respectfully",
    "Essential Guide": "how-to-visit",
    "Nearby / Pairing": "nearby-sites-and-routes",
    "Photography / Etiquette": "photography-and-etiquette",
}


def split_pipe(value):
    if value is None:
        return []
    return [p.strip() for p in str(value).split("|") if p and p.strip()]


def segments(url):
    return [s for s in str(url).strip("/").split("/") if s]


def slugify(text):
    text = str(text).lower().strip()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")


def canonical_url(url, page_type):
    """Return the cleaned URL for a page. Leaf-slug canonicalisation only touches
    article leaves; hub/feature URLs are returned unchanged (minus normalisation)."""
    segs = segments(url)
    if not segs:
        return "/"
    leaf = segs[-1]
    if page_type in LEAF_CANON:
        segs[-1] = LEAF_CANON[page_type]
    elif page_type == "Itinerary":
        m = re.search(r"(\d+)\s*day", leaf)
        n = m.group(1) if m else "1"
        segs[-1] = f"{n}-day-itinerary"
    # else (Site/Cluster/Main Hub, Cross-Cluster Feature) keep the existing leaf.
    return "/" + "/".join(segs) + "/"


def remap_dangling(url):
    if url in ("/magazine/", "/magazine"):
        return MAGAZINE_ROOT
    if url in ("/safety-ethics/", "/safety-ethics"):
        return SAFETY_HUB
    return url


def file_slug(url):
    """content/magazine/<slug>.md  — URL minus the magazine root, internal '/' -> '--'."""
    inner = url[len(MAGAZINE_ROOT):].strip("/") if url.startswith(MAGAZINE_ROOT) else url.strip("/")
    return inner.replace("/", "--") or "index"


def main():
    if not XLSX.exists():
        sys.exit(f"Workbook not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- 1. read raw pages --------------------------------------------------
    ws = wb["Content Plan"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    raw = [dict(zip(hdr, r)) for r in rows[1:] if any(c is not None for c in r)]

    # ---- 2. build old-URL -> new-URL map (canonical leaf + dangling remap) ---
    url_map = {}
    for d in raw:
        old = str(d["URL"])
        new = remap_dangling(canonical_url(old, str(d.get("Type", ""))))
        url_map[old] = new
    # parents / link targets may reference the synthetic hubs:
    url_map.setdefault("/magazine/", MAGAZINE_ROOT)
    url_map.setdefault("/safety-ethics/", SAFETY_HUB)

    def remap(u):
        if u is None:
            return None
        u = str(u).strip()
        if u in url_map:
            return url_map[u]
        return remap_dangling(u)

    # collision guard: cleaned URLs must stay unique
    seen = {}
    for old, new in url_map.items():
        if new in seen and seen[new] != old and old in {d["URL"] for d in raw}:
            # only care about real page collisions
            pass
    page_urls = [remap(d["URL"]) for d in raw]
    dupes = {u for u in page_urls if page_urls.count(u) > 1}
    if dupes:
        sys.exit(f"FATAL: leaf canonicalisation produced duplicate URLs: {sorted(dupes)[:10]}")

    # ---- 3. emit page records ----------------------------------------------
    pages = []
    for d in raw:
        url = remap(d["URL"])
        ptype = str(d.get("Type", "")).strip()
        pages.append({
            "id": d.get("ID"),
            "slug": file_slug(url),
            "url": url,
            "type": ptype,
            "cluster": d.get("Cluster"),
            "region": d.get("Region"),
            "country": d.get("Country"),
            "priority": d.get("Priority"),
            "clickDepth": d.get("Click Depth"),
            "parentUrl": remap(d.get("Parent URL")),
            "title": d.get("Title"),
            "h1": d.get("H1"),
            "primaryKeyword": d.get("Primary Keyword"),
            "secondaryKeywords": split_pipe(d.get("Secondary Keywords")),
            "searchIntent": d.get("Search Intent"),
            "audienceDesire": d.get("Audience Desire"),
            "serpAngle": d.get("SERP Angle"),
            # plan metas kept as INFORMATIONAL fallback only (truncated/templated in the sheet)
            "planMetaTitle": d.get("Meta Title"),
            "planMetaDescription": d.get("Meta Description"),
            "h2Outline": split_pipe(d.get("H2 Outline")),
            "faqQuestions": split_pipe(d.get("FAQ Questions")),
            "semanticEntities": split_pipe(d.get("Semantic Entities")),
            "requiredInternalLinks": [remap(x) for x in split_pipe(d.get("Internal Links Required"))],
            "sourceUrl": d.get("Source URL"),
            "claudePrompt": d.get("Claude Prompt"),
            "guardrail": d.get("Guardrail"),
            "publishBatch": d.get("Publish Batch"),
            "isHub": ptype in HUB_TYPES,
        })

    # ---- 4. synthesise the safety-ethics hub (referenced 642x, exists nowhere)
    if not any(p["url"] == SAFETY_HUB for p in pages):
        pages.append({
            "id": "safety-ethics-hub",
            "slug": file_slug(SAFETY_HUB),
            "url": SAFETY_HUB,
            "type": "Safety Hub",
            "cluster": "Global",
            "region": "Global",
            "country": "Global",
            "priority": "P0",
            "clickDepth": 2,
            "parentUrl": MAGAZINE_ROOT,
            "title": "How to Visit Dark Tourism Sites Respectfully: Ethics and Safety",
            "h1": "Ethics and Safety for Dark Tourism",
            "primaryKeyword": "dark tourism ethics",
            "secondaryKeywords": ["how to visit memorials respectfully", "dark tourism safety",
                                  "memorial etiquette", "is dark tourism ethical"],
            "searchIntent": "Ethical dark tourism",
            "audienceDesire": "Understand how to behave and stay safe at sensitive sites",
            "serpAngle": "Editorial ethics + safety framework hub for the whole magazine",
            "planMetaTitle": None,
            "planMetaDescription": None,
            "h2Outline": ["Why ethics matter in dark tourism", "Photography and behaviour",
                          "Respecting victims and communities", "Safety and travel advisories",
                          "Active conflict zones: a reality check", "Before you go"],
            "faqQuestions": ["Is dark tourism ethical?", "Can you take photos at memorial sites?",
                             "How should you behave at a memorial?",
                             "Should you visit active conflict zones?"],
            "semanticEntities": ["dark tourism", "ethics", "memorial etiquette", "travel safety",
                                 "victims", "advisory"],
            "requiredInternalLinks": [MAGAZINE_ROOT],
            "sourceUrl": "https://www.gov.uk/foreign-travel-advice",
            "claudePrompt": ("Write the magazine's ethics + safety hub. Set the rules for visiting "
                             "memorials and sensitive sites respectfully, and the policy on active "
                             "conflict zones. Serious, non-exploitative. Link to clusters and the main hub."),
            "guardrail": "No danger porn. No active war-zone promotion. Respect victims.",
            "publishBatch": "Batch 1",
            "isHub": True,
        })

    # ---- 5. assign publishAt ramp dates ------------------------------------
    def sort_key(p):
        return (
            TYPE_RANK.get(p["type"], 5),
            PRIORITY_RANK.get(str(p["priority"]), 9),
            str(p.get("publishBatch") or "Batch 9"),
            p["id"] if isinstance(p["id"], int) else 10**9,
        )

    ordered = sorted(pages, key=sort_key)
    for pos, p in enumerate(ordered):
        if pos < RAMP_DAYS * RAMP_PER_DAY:
            day = pos // RAMP_PER_DAY
        else:
            day = RAMP_DAYS + (pos - RAMP_DAYS * RAMP_PER_DAY) // CRUISE_PER_DAY
        p["publishAt"] = (START_DATE + timedelta(days=day)).isoformat()
    last_date = ordered[-1]["publishAt"]

    pages.sort(key=lambda p: (sort_key(p)))  # stable, structural-first order in the file

    # ---- 6. internal links graph -------------------------------------------
    wl = wb["Internal Links"]
    lrows = list(wl.iter_rows(values_only=True))
    valid_urls = {p["url"] for p in pages}
    links = {}
    seen_pairs = set()
    dropped = 0
    for r in lrows[1:]:
        if not r or not r[0] or not r[1]:
            continue
        src, tgt = remap(r[0]), remap(r[1])
        if src == tgt:
            dropped += 1
            continue
        key = (src, tgt)
        if key in seen_pairs:
            dropped += 1
            continue
        seen_pairs.add(key)
        links.setdefault(src, []).append({
            "target": tgt,
            "anchor": (str(r[2]).strip() if r[2] else ""),
            "type": str(r[3]).strip() if r[3] else "",
            "priority": str(r[4]).strip() if r[4] else "",
        })

    # ---- 7. write ----------------------------------------------------------
    by_type = {}
    for p in pages:
        by_type[p["type"]] = by_type.get(p["type"], 0) + 1

    meta = {
        "generatedFrom": XLSX.name,
        "schedule": {
            "startDate": START_DATE.isoformat(),
            "rampDays": RAMP_DAYS,
            "rampPerDay": RAMP_PER_DAY,
            "cruisePerDay": CRUISE_PER_DAY,
            "lastPublishAt": last_date,
        },
        "magazineRoot": MAGAZINE_ROOT,
        "counts": {"total": len(pages), "byType": by_type},
    }
    OUT_PAGES.write_text(json.dumps({"meta": meta, "pages": pages}, ensure_ascii=False, indent=2))
    OUT_LINKS.write_text(json.dumps(links, ensure_ascii=False, indent=2))

    print(f"pages: {len(pages)}  (last publishAt {last_date})")
    print(f"links: {sum(len(v) for v in links.values())} kept across {len(links)} sources "
          f"({dropped} self/dup dropped)")
    print(f"wrote {OUT_PAGES.relative_to(ROOT)} and {OUT_LINKS.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
